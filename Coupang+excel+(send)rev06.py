import openpyxl

# 박스히어로엑셀 파일 경로
boxhero_file_path = '박스히어로.xlsx'

# 쿠팡 커넥트 엑셀 파일 경로
coupang_file_path = '쿠팡커넥.xlsx'

# 박스히어로엑셀 파일 열기
boxhero_workbook = openpyxl.load_workbook(boxhero_file_path)
boxhero_sheet = boxhero_workbook.active

# 쿠팡 커넥트 엑셀 파일 열기
coupang_workbook = openpyxl.load_workbook(coupang_file_path)
coupang_sheet = coupang_workbook.active

# 박스히어로엑셀의 I열 값 읽기
boxhero_i_column = boxhero_sheet['I']
boxhero_b_column = boxhero_sheet['B']

# 쿠팡 커넥트 엑셀의 Q열 값 읽기
coupang_q_column = coupang_sheet['Q']
coupang_p_column = coupang_sheet['P']
coupang_b_column = coupang_sheet['B']

# ORDER1 엑셀 파일 열기
order1_workbook = openpyxl.load_workbook('order1.xlsx')
order1_sheet = order1_workbook.active

# 매칭된 것만 처리하기 위한 빈 리스트 생성
matched_rows = []

# 쿠팡 커넥트 엑셀 행 별로 매칭 여부 확인 및 P열, B열 값 출력
for i in range(2, len(coupang_q_column) + 1):
    coupang_value = coupang_q_column[i - 1].value
    matching_p_value = coupang_p_column[i - 1].value
    matching_b_value = coupang_b_column[i - 1].value

    if coupang_value is None or coupang_value == "":
        print(f"매칭되지 않거나 빈 셀인 쿠팡 커넥트 NO {i-1}.: P({matching_p_value}), B({matching_b_value})")
    else:
        matched = False

        for j in range(2, len(boxhero_i_column) + 1):
            boxhero_value = boxhero_i_column[j - 1].value
            if boxhero_value and boxhero_value in coupang_value:
                matched = True
                matched_rows.append(i)  # 매칭된 행 번호 저장
                break

#-----쿠팡커넥에 있는 내용이 order1 엑셀에 맵핑하여 복사
# 매칭된 행만 처리
for index, i in enumerate(matched_rows, start=2):
    # #----개수----
    # for W_row in coupang_sheet.iter_rows(min_row=2, max_row=i, min_col=23, max_col=23):
    #     for cell in W_row:
    #         order1_sheet['Q' + str(index)].value = cell.value

    # ----받는사람이름----
    for y_row in coupang_sheet.iter_rows(min_row=i, max_row=i, min_col=25, max_col=25):
        for cell in y_row:
            order1_sheet['C' + str(index)].value = cell.value  # index를 사용하여 순서대로 삽입
    # ----주문번호----
    for C_row in coupang_sheet.iter_rows(min_row=2, max_row=i, min_col=3, max_col=3):
        for cell in C_row:
            order1_sheet['F' + str(index)].value = cell.value

    # ----전화번호----
    for AK_row in coupang_sheet.iter_rows(min_row=2, max_row=i, min_col=37, max_col=37):
        for cell in AK_row:
            order1_sheet['J' + str(index)].value = cell.value
    # ----전화번호2----
    for Z_row in coupang_sheet.iter_rows(min_row=2, max_row=i, min_col=26,max_col=26):
        for cell in Z_row:
            order1_sheet['k' + str(index)].value = cell.value

    # ----Personal Customs Clearance Code (PCCC)----
    for AJ_row in coupang_sheet.iter_rows(min_row=2, max_row=i, min_col=36, max_col=36):
        for cell in AJ_row:
            order1_sheet['I' + str(index)].value = cell.value

    # #----상품명----
    # for P_row in coupang_sheet.iter_rows(min_row=2, max_row=i, min_col=16, max_col=16):
    #     for cell in P_row:
    #         order1_sheet['P' + str(index)].value = cell.value

    # ----우편번호----
    for AC_row in coupang_sheet.iter_rows(min_row=2, max_row=i, min_col=29, max_col=29):
        for cell in AC_row:
            order1_sheet['L' + str(index)].value = cell.value


    #----집주소----
    for AD_row in coupang_sheet.iter_rows(min_row=2, max_row=i, min_col=30, max_col=30):
        for cell in AD_row:
            order1_sheet['M' + str(index)].value = cell.value

    #----메모----
    for AE_row in coupang_sheet.iter_rows(min_row=2, max_row=i, min_col=31, max_col=31):
        for cell in AE_row:
            order1_sheet['O' + str(index)].value = cell.value
    #----단가----
    for S_row in coupang_sheet.iter_rows(min_row=2, max_row=i, min_col=19, max_col=19):
        for cell in S_row:
            order1_sheet['S' + str(index)].value = cell.value


#-----박스히어로에 있는 내용이 order1 엑셀에 맵핑하여 복사


# 박스히어로 엑셀의 I열 값 읽기
boxhero_i_column = boxhero_sheet['I']
boxhero_m_column = boxhero_sheet['M']  # 박스히어로의 M열 추가

# 쿠팡 커넥트 엑셀의 Q열 값 읽기
coupang_q_column = coupang_sheet['Q']

# 매칭된 행 정보를 저장할 리스트
matched_rows = []

# 박스히어로의 N열 값 읽기
boxhero_n_column = boxhero_sheet['N']
boxhero_p_column = boxhero_sheet['P']
boxhero_b_column = boxhero_sheet['B']

# 쿠팡커넥의 Q 열과 박스히어로의 I열 맵핑
for i in range(2, len(coupang_q_column) + 1):
    coupang_value = coupang_q_column[i - 1].value

    if coupang_value is not None and coupang_value != "":
        matched_m_value = None
        matched_n_value = None
        matched_p_value = None
        matched_b_value = None

        for j in range(2, len(boxhero_i_column) + 1):
            boxhero_value = boxhero_i_column[j - 1].value

            if boxhero_value and boxhero_value in coupang_value:
                matched_m_value = boxhero_m_column[j - 1].value
                matched_n_value = boxhero_n_column[j - 1].value
                matched_p_value = boxhero_p_column[j - 1].value
                matched_b_value = boxhero_b_column[j - 1].value
                break

        if matched_m_value is not None:
            matched_rows.append((matched_m_value, matched_n_value, matched_p_value, matched_b_value))

# HS CODE 및 목록일반
for index, (m_value, n_value, p_value, b_value) in enumerate(matched_rows, start=2):
    order1_sheet.cell(row=index, column=18, value=m_value)  # HS CODE 열
    order1_sheet.cell(row=index, column=7, value=n_value)  # 목록일반 열
    order1_sheet.cell(row=index, column=4, value=p_value)  # 무게
    order1_sheet.cell(row=index, column=16, value=b_value)  # 영문상품명

# A열에 번호 매기기
for index, i in enumerate(matched_rows, start=1):
    order1_sheet.cell(row=index + 1, column=1, value=index)


# ORDER1 엑셀 파일 저장
order1_workbook.save('order1.xlsx')


# 박스히어로 엑셀의 I열 값 읽기
boxhero_i_column = boxhero_sheet['I']

# 쿠팡 커넥트 엑셀의 Q열 값 읽기
coupang_q_column = coupang_sheet['Q']

# ORDER1 엑셀 파일 열기
order1_workbook = openpyxl.load_workbook('order1.xlsx')
order1_sheet = order1_workbook.active

# 매칭된 것만 처리하기 위한 빈 리스트 생성
matched_rows = []

# 쿠팡 커넥트 엑셀 행 별로 매칭 여부 확인 및 P열, B열 값 출력
for i in range(2, len(coupang_q_column) + 1):
    coupang_value = coupang_q_column[i - 1].value

    if coupang_value is None or coupang_value == "":
        # print(f"매칭되지 않거나 빈 셀인 쿠팡 커넥트 {i-1}행")  # 해당 프린트문 주석 처리
        pass  # 해당 라인을 무시하도록 pass 문 사용
    else:
        matched = False

        for j in range(2, len(boxhero_i_column) + 1):
            boxhero_value = boxhero_i_column[j - 1].value
            if boxhero_value and boxhero_value in coupang_value:
                matched = True
                matched_rows.append(i)  # 매칭된 행 번호 저장
                break


#-----쿠팡커넥에 있는 내용이 order1 엑셀에 맵핑하여 복사
# 매칭된 행만 처리
for index, i in enumerate(matched_rows, start=2):
    # 개수 코드(W)에서 S 뒤에 있는 숫자만큼 개수를 추가
    coupang_qty = coupang_sheet.cell(row=i, column=23).value
    product_code = coupang_sheet.cell(row=i, column=17).value

    if coupang_qty and product_code:
        # 프로덕트 코드 끝이 'S숫자' 형식인 경우 숫자를 추출
        last_part = product_code.split('S')[-1]
        if last_part.isdigit():
            try:
                s_qty = int(last_part)
                coupang_qty = int(coupang_qty)  # coupang_qty를 숫자로 형변환
                order1_sheet.cell(row=index, column=17, value=coupang_qty * s_qty)
            except ValueError:
                pass
        else:
            # S + '숫자' 형식이 아닌 경우 그냥 coupang_qty를 넣음
            order1_sheet.cell(row=index, column=17, value=coupang_qty)

# ORDER1 엑셀 파일 저장
order1_workbook.save('order1.xlsx')


